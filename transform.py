from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import os
from dotenv import load_dotenv
import yagmail

load_dotenv()

MAIL_AUTOR = os.getenv("MAIL_AUTOR")
APP_GMAIL_PASS = os.getenv("APP_GMAIL_PASS")
MAIL_DESTINO = os.getenv("MAIL_DESTINO")

today = datetime.now()

def export_excel(data_docs, data_docs_2, data_prest):
    
  # Hoja 1 - Resumen general

  wb = Workbook()
  ws = wb.active
  ws.title = "Resumen general 2025"

  headers_resumen_2025 = ["PRESTACION ID", "ALUMNO ID", "NOMBRE", "DNI", "OS", "PREST. SAIE", 
                      "CRED. DNI", "CRED. OS", "CRED. CUD", "AD", "ORD. MED.", "RHC", 
                      "PLAN TR.", "PRESUP.", "OTROS", "INF. ADM.", "INF. INIC. TER.",
                      "INF. SEMEST.", "INF. FINAL"]
  
  ws.append(headers_resumen_2025)

  for cell in ws[1]:
      cell.font = Font(bold=True)

  for row in data_docs:
      ws.append(row)

  # Hoja 2 - Resumen general 2026

  ws2 = wb.create_sheet(title="Resumen general 2026")

  headers_resumen_2026 = ["PRESTACION ID", "ALUMNO ID", "NOMBRE", "DNI", "OS", 
                          "AD", "ORD. MED.", "RHC", "PLAN TR.", "PRESUP.", 
                          "OTROS", "INF. ADM.", "INF. INIC. TER.","INF. SEMEST.", 
                          "INF. FINAL"]

  ws2.append(headers_resumen_2026)

  for cell in ws2[1]:
      cell.font = Font(bold=True)

  for row in data_docs_2:
      ws2.append(row)

  # Hoja 3 - Altas y bajas de prestaciones

  ws3 = wb.create_sheet(title="Altas y bajas de prest.")

  headers_altas_bajas = ["AÑO", "MES", "ALTAS", "BAJAS"]

  ws3.append(headers_altas_bajas)

  for cell in ws3[1]:
      cell.font = Font(bold=True)

  for row in data_prest:
      ws3.append(row)

  nombre_archivo = f"reporte_terapias_{today.strftime('%Y-%m-%d')}.xlsx"
  wb.save(nombre_archivo)
  print(f"Archivo Excel generado: {nombre_archivo}")
  return nombre_archivo

def enviar_correo(nombre_archivo):
  try:
    yag = yagmail.SMTP(MAIL_AUTOR, APP_GMAIL_PASS)
    yag.send(
      to=MAIL_DESTINO,
      subject="Reporte general de Terapias",
      contents= """Buenos días, se adjunta el reporte semanal del área de Terapias.
              \nSaludos,\nMariano López - Ailes Inclusión.""",
      attachments=nombre_archivo
    )
    print("Correo enviado correctamente.")
  except Exception as e:
    print("Error al enviar el correo:", e)